from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import os
import re
from openpyxl import load_workbook
import pandas as pd

from drive_utils import (
    discover_service_account_file,
    download_file_bytes,
    find_child_folder_id,
    list_files_in_folder_tree,
    upload_or_replace_file,
)
from master_shared import (
    calc_achieve as shared_calc_achieve,
    calc_status as shared_calc_status,
    default_input_folder,
    default_output_file,
    project_health as shared_project_health,
    safe_pct as shared_safe_pct,
)

# ==========================================================
# CONFIG
# ==========================================================

INPUT_FOLDER = default_input_folder()

OUTPUT_FILE = default_output_file()

SHEET_NAME = "Activity Plan"

DRIVE_ROOT_FOLDER = os.getenv("GOOGLE_DRIVE_ROOT_FOLDER", "11YZWsFx4MDk0ejyp2EbhfAnU5zgv9ZeP")
DRIVE_RAW_FOLDER_NAME = "raw"
DRIVE_MAIN_FOLDER_NAME = "main"

# ==========================================================
# HELPER
# ==========================================================

def safe_pct(value):
    return shared_safe_pct(value)


def calc_achieve(plan, actual):
    return shared_calc_achieve(plan, actual)


def calc_status(achieve):
    return shared_calc_status(achieve)


@dataclass(frozen=True)
class InputSource:
    name: str
    stem: str
    path: Path | None = None
    file_bytes: bytes | None = None


def parse_project_filename(filename: str) -> tuple[str | None, str | None, str | None]:
    stem = Path(filename).stem
    match = re.match(r"^CAM_(?P<department>[A-Za-z0-9]+)_(?P<project_id>.+)$", stem)
    if not match:
        return None, None, None

    return "CAM", match.group("department").upper(), match.group("project_id").upper()


def _build_input_sources() -> tuple[list[InputSource], str | None, str | None]:
    try:
        credentials_file = discover_service_account_file()
        raw_folder_id = find_child_folder_id(
            credentials_file,
            DRIVE_ROOT_FOLDER,
            DRIVE_RAW_FOLDER_NAME,
        ) or DRIVE_ROOT_FOLDER
        drive_files = list_files_in_folder_tree(credentials_file, raw_folder_id)
        if drive_files:
            print("=" * 60)
            print("GOOGLE DRIVE RAW FILES FOUND")
            print("=" * 60)
            sources = []
            for item in drive_files:
                print(item.name)
                file_bytes = download_file_bytes(credentials_file, item.file_id)
                sources.append(
                    InputSource(
                        name=item.name,
                        stem=Path(item.name).stem,
                        file_bytes=file_bytes,
                    )
                )
            print()
            print("TOTAL FILE :", len(sources))
            print()
            main_folder_id = find_child_folder_id(
                credentials_file,
                DRIVE_ROOT_FOLDER,
                DRIVE_MAIN_FOLDER_NAME,
            ) or DRIVE_ROOT_FOLDER
            return sources, str(credentials_file), main_folder_id
    except Exception as exc:
        print(f"GOOGLE DRIVE MODE DISABLED: {exc}")

    files = sorted(INPUT_FOLDER.glob("*.xlsx"))
    print("=" * 60)
    print("FILES FOUND")
    print("=" * 60)
    for f in files:
        print(f.name)
    print()
    print("TOTAL FILE :", len(files))
    print()
    return (
        [
            InputSource(
                name=f.name,
                stem=f.stem,
                path=f,
            )
            for f in files
        ],
        None,
        None,
    )

# ==========================================================
# STORAGE
# ==========================================================

project_records = []
progress_records = []

# ==========================================================
# FILE LIST
# ==========================================================

input_sources, drive_credentials_path, drive_main_folder_id = _build_input_sources()

# ==========================================================
# LOOP FILE
# ==========================================================

for file_source in input_sources:

    try:

        print(f"\nPROCESSING : {file_source.name}")

        # ==================================================
        # FILE INFO
        # ==================================================

        site, department, project_id = parse_project_filename(file_source.name)
        project_key = None if department is None or project_id is None else f"{department}_{project_id}"

        row_added = False
        if site is not None and department is not None and project_id is not None:
            project_records.append({
                "SITE": site,
                "DEPARTMENT": department,
                "PROJECT_ID": project_id,
                "PROJECT_KEY": project_key,
                "FILE_NAME": file_source.name
            })
            row_added = True

        print(
            "DEBUG PROJECT_MASTER | "
            f"filename={file_source.name} | "
            f"department={department} | "
            f"project={project_id} | "
            f"added={row_added}"
        )

        # ==================================================
        # OPEN FILE
        # ==================================================

        wb_source = (
            BytesIO(file_source.file_bytes)
            if file_source.file_bytes is not None
            else file_source.path
        )

        wb = load_workbook(
            wb_source,
            data_only=True
        )

        if SHEET_NAME not in wb.sheetnames:

            print(
                f"Sheet {SHEET_NAME} not found"
            )

            continue

        ws = wb[SHEET_NAME]

        # ==================================================
        # DATE HEADER
        # ==================================================

        timeline_dates = []

        for col in range(19, 109):

            timeline_dates.append(
                ws.cell(7, col).value
            )

        # ==================================================
        # LOOP ACTION PLAN
        # ==================================================

        row = 8

        while True:

            no = ws.cell(row, 2).value

            if no is None:
                break

            factor = ws.cell(row, 3).value
            problem = ws.cell(row, 4).value
            category = ws.cell(row, 5).value
            action_plan = ws.cell(row, 6).value
            pic = ws.cell(row, 7).value

            plan_start = ws.cell(row, 8).value
            due_date = ws.cell(row, 9).value

            for idx in range(len(timeline_dates)):

                current_date = timeline_dates[idx]

                plan_col = 19 + idx
                actual_col = 112 + idx

                plan_pct = safe_pct(
                    ws.cell(
                        row,
                        plan_col
                    ).value
                )

                actual_pct = safe_pct(
                    ws.cell(
                        row,
                        actual_col
                    ).value
                )

                if (
                    plan_pct is None
                    and actual_pct is None
                ):
                    continue

                achieve_pct = calc_achieve(
                    plan_pct,
                    actual_pct
                )

                status = calc_status(
                    achieve_pct
                )

                progress_records.append({

                    "SITE": site,
                    "DEPARTMENT": department,
                    "PROJECT_ID": project_id,

                    "NO": no,

                    "FACTOR": factor,
                    "PROBLEM": problem,
                    "CATEGORY": category,

                    "ACTION_PLAN": action_plan,
                    "PIC": pic,

                    "PLAN_START": plan_start,
                    "DUE_DATE": due_date,

                    "DATE": current_date,

                    "PLAN_PCT": plan_pct,
                    "ACTUAL_PCT": actual_pct,
                    "ACHIEVE_PCT": achieve_pct,

                    "STATUS": status
                })

            row += 1

    except Exception as e:

        print(
            f"ERROR : {file_source.name}"
        )

        print(e)

# ==========================================================
# DATAFRAME
# ==========================================================

project_df = pd.DataFrame(
    project_records
)

progress_df = pd.DataFrame(
    progress_records
)

# ==========================================================
# PROJECT DAILY PROGRESS
# ==========================================================

daily_records = []

group_cols = [
    "SITE",
    "DEPARTMENT",
    "PROJECT_ID",
    "DATE"
]

for keys, group in progress_df.groupby(group_cols):

    site, department, project_id, date = keys

    total_activity = group["NO"].nunique()

    active_group = group[
        group["PLAN_PCT"].notna()
    ]

    active_activity = (
        active_group["NO"]
        .nunique()
    )

    if active_activity == 0:
        continue

    plan_project = round(
        active_group["PLAN_PCT"].mean(),
        2
    )

    actual_project = round(
        active_group["ACTUAL_PCT"].mean(),
        2
    )

    achieve_project = None

    if (
        plan_project is not None
        and plan_project > 0
    ):

        achieve_project = (
            actual_project
            / plan_project
        ) * 100

        achieve_project = min(
            round(
                achieve_project,
                2
            ),
            120
        )

    status = None

    if achieve_project is not None:

        if achieve_project >= 100:
            status = "ON TRACK"

        elif achieve_project >= 90:
            status = "WARNING"

        else:
            status = "DELAY"

    daily_records.append({

        "SITE": site,

        "DEPARTMENT": department,

        "PROJECT_ID": project_id,

        "DATE": date,

        "TOTAL_ACTIVITY": total_activity,

        "ACTIVE_ACTIVITY": active_activity,

        "PLAN_PROJECT": plan_project,

        "ACTUAL_PROJECT": actual_project,

        "ACHIEVE_PROJECT": achieve_project,

        "STATUS": status

    })

project_daily_df = pd.DataFrame(
    daily_records
)

# ==========================================================
# SUMMARY PROJECT
# ==========================================================

summary_project_df = (

    project_daily_df

    .groupby(
        [
            "SITE",
            "DEPARTMENT",
            "PROJECT_ID"
        ],
        as_index=False
    )

    .agg({

        "TOTAL_ACTIVITY": "max",

        "PLAN_PROJECT": "last",

        "ACTUAL_PROJECT": "last",

        "ACHIEVE_PROJECT": "last"

    })

)

summary_project_df["STATUS"] = (
    summary_project_df[
        "ACHIEVE_PROJECT"
    ]
    .apply(
        lambda x:
        "ON TRACK"
        if pd.notna(x) and x >= 100
        else (
            "WARNING"
            if pd.notna(x) and x >= 90
            else "DELAY"
        )
    )
)

# ==========================================================
# VALIDATION
# ==========================================================

print()
print("=" * 60)
print("PROJECT")
print(len(project_df))

print("PROGRESS")
print(len(progress_df))
print("=" * 60)

if progress_df.empty:

    raise ValueError(
        "Tidak ada data berhasil dibaca."
    )

# ==========================================================
# DATE CLEANING
# ==========================================================

for col in [
    "DATE",
    "PLAN_START",
    "DUE_DATE"
]:

    progress_df[col] = pd.to_datetime(
        progress_df[col],
        errors="coerce"
    )

# ==========================================================
# SUMMARY
# ==========================================================

summary_df = (

    progress_df

    .groupby(
        "DEPARTMENT",
        as_index=False
    )

    .agg({

        "PROJECT_ID": "nunique",
        "NO": "count",
        "ACHIEVE_PCT": "mean"

    })

)

summary_df.columns = [

    "DEPARTMENT",
    "TOTAL_PROJECT",
    "TOTAL_RECORD",
    "AVG_ACHIEVE"

]

summary_df["AVG_ACHIEVE"] = (
    summary_df["AVG_ACHIEVE"]
    .round(2)
)

# ==========================================================
# LATEST STATUS
# ==========================================================

latest_df = (

    progress_df

    .sort_values(
        "DATE"
    )

    .groupby(
        [
            "SITE",
            "DEPARTMENT",
            "PROJECT_ID",
            "NO"
        ]
    )

    .tail(1)

)

# ==========================================================
# IMPROVE
# ==========================================================
executive_summary_df = pd.DataFrame({

    "KPI": [

        "Total Department",
        "Total Project",
        "Total Action Plan",
        "Average Achievement"

    ],

    "VALUE": [

        progress_df["DEPARTMENT"].nunique(),

        progress_df["PROJECT_ID"].nunique(),

        progress_df["NO"].nunique(),

        round(
            progress_df["ACHIEVE_PCT"].mean(),
            2
        )

    ]

})

department_performance_df = (

    latest_df

    .groupby(
        "DEPARTMENT",
        as_index=False
    )

    .agg(

        TOTAL_PROJECT=(
            "PROJECT_ID",
            "nunique"
        ),

        TOTAL_ACTION=(
            "NO",
            "count"
        ),

        AVG_ACHIEVE=(
            "ACHIEVE_PCT",
            "mean"
        )

    )

)

department_performance_df["AVG_ACHIEVE"] = (
    department_performance_df["AVG_ACHIEVE"]
    .round(2)
)

pic_performance_df = (

    latest_df

    .groupby(
        "PIC",
        as_index=False
    )

    .agg(

        TOTAL_ACTION=(
            "NO",
            "count"
        ),

        AVG_ACHIEVE=(
            "ACHIEVE_PCT",
            "mean"
        )

    )

)

pic_performance_df["AVG_ACHIEVE"] = (
    pic_performance_df["AVG_ACHIEVE"]
    .round(2)
)

pic_performance_df = (
    pic_performance_df
    .sort_values(
        "AVG_ACHIEVE",
        ascending=False
    )
)

factor_performance_df = (

    latest_df

    .groupby(
        "FACTOR",
        as_index=False
    )

    .agg(

        TOTAL_ACTION=(
            "NO",
            "count"
        ),

        AVG_ACHIEVE=(
            "ACHIEVE_PCT",
            "mean"
        )

    )

)

factor_performance_df["AVG_ACHIEVE"] = (
    factor_performance_df["AVG_ACHIEVE"]
    .round(2)
)

category_performance_df = (

    latest_df

    .groupby(
        "CATEGORY",
        as_index=False
    )

    .agg(

        TOTAL_ACTION=(
            "NO",
            "count"
        ),

        AVG_ACHIEVE=(
            "ACHIEVE_PCT",
            "mean"
        )

    )

)

category_performance_df["AVG_ACHIEVE"] = (
    category_performance_df["AVG_ACHIEVE"]
    .round(2)
)

delay_analysis_df = (

    progress_df[
        progress_df["STATUS"] == "DELAY"
    ]

    .groupby(

        [
            "SITE",
            "DEPARTMENT",
            "PROJECT_ID",
            "NO",
            "ACTION_PLAN",
            "PIC"
        ],

        as_index=False

    )

    .size()

)

delay_analysis_df.columns = [

    "SITE",
    "DEPARTMENT",
    "PROJECT_ID",
    "NO",
    "ACTION_PLAN",
    "PIC",
    "DELAY_DAYS"

]

delay_analysis_df = (

    delay_analysis_df

    .sort_values(
        "DELAY_DAYS",
        ascending=False
    )

)

project_health_df = (

    summary_project_df

    .copy()

)

def project_health(score):
    return shared_project_health(score)

project_health_df["HEALTH"] = (
    project_health_df["ACHIEVE_PROJECT"]
    .apply(project_health)
)

project_ranking_df = (

    summary_project_df

    .sort_values(
        "ACHIEVE_PROJECT",
        ascending=False
    )

    .reset_index(
        drop=True
    )

)

project_ranking_df.insert(
    0,
    "RANK",
    range(
        1,
        len(project_ranking_df) + 1
    )
)

print("SUMMARY_PROJECT :", len(summary_project_df))
print("EXECUTIVE_SUMMARY :", len(executive_summary_df))
print("DEPARTMENT_PERFORMANCE :", len(department_performance_df))
print("PIC_PERFORMANCE :", len(pic_performance_df))
print("FACTOR_PERFORMANCE :", len(factor_performance_df))
print("CATEGORY_PERFORMANCE :", len(category_performance_df))
print("DELAY_ANALYSIS :", len(delay_analysis_df))
print("PROJECT_HEALTH :", len(project_health_df))
print("PROJECT_RANKING :", len(project_ranking_df))

# ==========================================================
# EXPORT
# ==========================================================


OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    project_df.to_excel(
        writer,
        sheet_name="PROJECT_MASTER",
        index=False
    )

    progress_df.to_excel(
        writer,
        sheet_name="PROGRESS_MASTER",
        index=False
    )

    summary_df.to_excel(
        writer,
        sheet_name="SUMMARY",
        index=False
    )

    latest_df.to_excel(
        writer,
        sheet_name="LATEST_STATUS",
        index=False
    )

    project_daily_df.to_excel(
        writer,
        sheet_name="PROJECT_DAILY_PROGRESS",
        index=False
    )

    summary_project_df.to_excel(
        writer,
        sheet_name="SUMMARY_PROJECT",
        index=False
    )

    executive_summary_df.to_excel(
        writer,
        sheet_name="EXECUTIVE_SUMMARY",
        index=False
    )

    department_performance_df.to_excel(
        writer,
        sheet_name="DEPARTMENT_PERFORMANCE",
        index=False
    )

    pic_performance_df.to_excel(
        writer,
        sheet_name="PIC_PERFORMANCE",
        index=False
    )

    factor_performance_df.to_excel(
        writer,
        sheet_name="FACTOR_PERFORMANCE",
        index=False
    )

    category_performance_df.to_excel(
        writer,
        sheet_name="CATEGORY_PERFORMANCE",
        index=False
    )

    delay_analysis_df.to_excel(
        writer,
        sheet_name="DELAY_ANALYSIS",
        index=False
    )

    project_health_df.to_excel(
        writer,
        sheet_name="PROJECT_HEALTH",
        index=False
    )

    project_ranking_df.to_excel(
        writer,
        sheet_name="PROJECT_RANKING",
        index=False
    )

print("MASTER workbook generated:", OUTPUT_FILE)

if drive_credentials_path and drive_main_folder_id:
    try:
        uploaded = upload_or_replace_file(
            drive_credentials_path,
            drive_main_folder_id,
            OUTPUT_FILE.name,
            OUTPUT_FILE.read_bytes(),
        )
        print("MASTER workbook uploaded to Google Drive:", uploaded.name)
    except Exception as exc:
        print("WARNING: failed to upload MASTER workbook to Google Drive")
        print(exc)
